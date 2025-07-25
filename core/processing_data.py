from pathlib import Path

import pandas as pd  # type: ignore
from openpyxl import load_workbook  # type: ignore
from openpyxl.styles import numbers  # type: ignore # noqa: F401

from utils.config_logger import setup_logger

pd.options.mode.copy_on_write = True 

logger = setup_logger()

class ExcelProcessor:
    def __init__(self, file_path: Path):
        self.file_path = file_path

    def check_df_empty(self, df: pd.DataFrame) -> bool:
        """
        Verifica se o DataFrame está vazio.

        Args:
            df (pd.DataFrame): O DataFrame a ser verificado.

        Returns:
            bool: True se o DataFrame estiver vazio, False caso contrário.
        """
        return df.empty    

    def load_excel_to_df(self, sheet_name: str | None = 'Sheet1', columns_to_use: list[ str | int ] | None = None) -> pd.DataFrame:
        """
        Carrega dados de um arquivo Excel para Pandas DataFrame.

        Args:
            sheet_name (str | None): O nome da planilha a ser carregado. Se "None" a primeira planilha é carregada
            columns_to_use (list[int | str] | None): Uma lista de indices ou nomes das colunas a serem usadas

        Returns:
            pd.DataFrame: O DataFrame carregado

        Raises:
            FileNotFoundError: Se o arquivo excel especificado não existir.
            ValueError: Se o nome da planilha ou colunas são invalidas.
        """       
        if not self.file_path.exists():
            logger.error(f'Excel file not found at: {self.file_path}')
            raise FileNotFoundError(f'The file {self.file_path} does not exist')
        try:
            return  pd.read_excel(
                io=self.file_path,
                sheet_name=sheet_name,
                usecols=columns_to_use,
                engine='openpyxl'
            ) # type: ignore
        except Exception as e:
            logger.error(f' Erros loading Excel file {self.file_path} : {e}')
            raise ValueError(f'Could not load data from {self.file_path}. Check sheet_name or columns')  # noqa: B904

    def view_df(self, df: pd.DataFrame, name_columns_load: list[ str | int ]) -> None:
        """
        Mostra o Dataframe com as colunas especificadas

        Args:
            df (pd.DataFrame): O DataFrame de entrada
            name_columns_load (list[str|int]): Lista com os nomes das colunas ou indices para mostrar
        Returns:
            None

        Raise:
            KeyErro: Se a coluna especifica não existir no DataFrame.
            ValueError: Se o DataFrame estiver vazio.
        """
        if self.check_df_empty(df):
            logger.error("The DataFrame is empty. Cannot display columns.")
            raise ValueError("The DataFrame is empty. Please provide a valid DataFrame.")
        
        for col in name_columns_load:
            if col not in df.columns:
                logger.warning(f"Column '{col}' not found in DataFrame.")
                raise KeyError(f"The columns {col} not found in DataFrame. Check name's columns")
                
        logger.info(df[name_columns_load])
    
    def filter_open_invoices(self, df: pd.DataFrame, type_doc_column: str, exclude_value: str, account_column:str, account_code: str) -> pd.DataFrame:
        """
        Filtra o DataFrame para incluir apenas faturas em aberto com base no tipo doc e numero conta contabil

        Args:
            df (pd.DataFrame): O DataFrame de entrada
            type_doc_column (str): Nome da coluna indicando o Tipo da fatura
            excluded_status_value (str): O valor a ser excluida da coluna (ex.:, 'Contas a Pagar').
            account_column (str): O nome da coluna com as contas contabeis (ex.:, 'Conta').
            account_number (str): O numero especifico da conta contabil.

        Returns:
            pd.DataFrame: Um DataFrame contendo apenas as notas em aberto

        Raises:
            KeyError: Se a coluna especifica não existir no DataFrame.
            ValueError: Se o DataFrame estiver vazio.
        """
        if self.check_df_empty(df):
            logger.error("The DataFrame is empty. Cannot filter open invoices.")
            raise ValueError("The DataFrame is empty. Please provide a valid DataFrame.")

        if type_doc_column not in df.columns:
            logger.error(f' Status colunm {type_doc_column} not found in DataFrame')
            raise KeyError(f'Column {type_doc_column} is missing from the Dataframe')
        
        if account_column not in df.columns:
            logger.error(f' Account code {account_column} not found in DataFrame')
            raise KeyError(f'Column {account_column} is missing from the Dataframe')
        
        open_invoices = df[~df[type_doc_column].isin([exclude_value])].copy()
        filtered_df = open_invoices[open_invoices[account_column]. isin([account_code])].copy()
        return filtered_df # type: ignore

    def reconcile_invoices(self, invoices_to_pay_df: pd.DataFrame, paid_invoices_df: pd.DataFrame) -> pd.DataFrame:
        """
        Concilia faturas a pagar com faturas já pagas.

        Args:
            invoices_to_pay_df (pd.DataFrame): DataFrame de faturas que precisam ser pagas.
            paid_invoices_df (pd.DataFrame): DataFrame de faturas já pagas.

        Returns:
            pd.DataFrame: Um DataFrame de faturas que estão abertas e precisam ser pagas.

        Raises:
            ValueError: Se a coluna 'No Doc SAP' estiver faltando em um dos dois DataFrames .
        """
        if "No Doc SAP" not in invoices_to_pay_df.columns or "No Doc SAP" not in paid_invoices_df.columns:
            logger.error('Missing [No Doc SAP] columns for invoice reconciliation')
            raise ValueError('Both DataFrame must contain a [No Doc SAP] column for reconciliation')
        
        if self.check_df_empty(invoices_to_pay_df) or self.check_df_empty(paid_invoices_df):
            logger.error("One or both DataFrames are empty. Cannot reconcile invoices.")
            raise ValueError("Both DataFrames must contain data for reconciliation.")

        open_for_payment = invoices_to_pay_df[~invoices_to_pay_df["No Doc SAP"].isin(paid_invoices_df["No Doc SAP"])].copy()
        return open_for_payment # type: ignore

    def convert_date_df(self, df: pd.DataFrame, date_columns: list[ int | str ]) -> pd.DataFrame:
        """
        Converte as colunas especificadas do DataFrame em datetime

        Args:
            df (pd.DataFrame): O DataFrame de entrada
            date_columns (list[str|int]): Lista com os nomes das colunas ou indices para serem convertidas
        Returns:
            DataFrame com as colunas de entrada convertidas em datetime

        Raise:
            KeyError: Se a coluna especifica não existir no DataFrame.
        """
        df = df.copy()
        date_columns = [col.strip() for col in date_columns] # type: ignore
        for col in date_columns:
            if col not in df.columns:
                logger.warning(f"Column '{col}' not found in DataFrame")
                raise KeyError(f"The columns {col} not found in DataFrame. Check name's columns")
        
        for col in date_columns:        
            try:                      
                df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce').copy() # type: ignore
                logger.debug(f"Type before convertion: {df[col].dtype}")
                if df[col].isnull().any():
                    logger.warning(f"Some values in column '{col}' could not be converted to datetime and are set to NaT.")
            except Exception as e:
                logger.error(f'Could not convert column {col} to datetime. {e}')        
        return df
    
    def convert_values_df(self, df: pd.DataFrame, value_columns: str) -> pd.DataFrame:
        df = df.copy()  
        if value_columns in df.columns:  
            try:
                if pd.api.types.is_string_dtype(df[value_columns]):
                    df[value_columns] = (
                        df[value_columns]
                        .str.replace('.', '', regex=False)
                        .str.replace(',', '.', regex=False)
                    )
                df[value_columns] = pd.to_numeric(df[value_columns], errors='coerce')
                logger.debug(f"Type before convertion: {df[value_columns].dtype}")

            except Exception as e:
                logger.error(f"Could not convert column '{value_columns}' to numeric. Error: {e}")
        else:
            logger.warning(f"Column '{value_columns}' not found in DataFrame for value conversion")
        return df
    
    def filter_by_due_date(self, df: pd.DataFrame, due_data_column: str, target_date: str) -> pd.DataFrame:
        """
        Filtra faturas por uma determinada data de vencimento. Garante que a coluna de data esteja no formato data e hora.

        Args:
            df (pd.DataFrame): The input DataFrame of invoices.
            due_data_column (str): O nome da coluna de data de vencimento (ex:., 'Data Vencimento' or 'Dt Vencimento').
            target_date (str):  A string de data limite  (ex:., 'DD/MM/YYYY').

        Returns:
            pd.DataFrame:  Um DataFrame ordenado pela data de vencimento e filtrado pela data de destino

        Raises:
            ValueError: I: Se a análise de data falhar ou se as colunas obrigatórias estiverem ausentes
        """
        if due_data_column not in df.columns:
            logger.error(f"Missing required date column: '{due_data_column}'.")
            raise ValueError(f"Required column '{due_data_column}' must be present in the DataFrame.")

        try:
            
            if not pd.api.types.is_datetime64_any_dtype(df[due_data_column]):
                df.loc[:, due_data_column] = pd.to_datetime(df[due_data_column], dayfirst=True, errors='coerce')
                logger.debug(f"Tipo após conversão: {df[due_data_column].dtype}")

            target_dt = pd.to_datetime(target_date, dayfirst=True)
            
        except Exception as e:
            logger.error(f"Error converting date columns or target date in filter_by_due_date: {e}")
            raise ValueError(f"Invalid date format or column for filtering. Error: {e}")  # noqa: B904

        logger.info(f'Current size df {len(df)}')
        df_cleaned = df.dropna(subset=[due_data_column]).copy()
        logger.info(f'Size of DataFrame before operation {len(df)}')

        df_filtered = df_cleaned[df_cleaned[due_data_column] <= target_dt].copy()
        
        return df_filtered.sort_values(by=due_data_column) # type: ignore

    def concatenate_dfs(self, df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
        """
        Concatena dois DataFrames.

        Args:
            df1 (pd.DataFrame): O primeiro DataFrame.
            df2 (pd.DataFrame): O segundo DataFrame.

        Returns:
            pd.DataFrame: Um DataFrame concatenado.
        """
        return pd.concat([df1, df2], ignore_index=True)


class ExcelFormater:
    def __init__(self, file_path: Path, df_to_export_excel: pd.DataFrame) -> None:
        self.file_path = file_path
        self.df_to_export_excel = df_to_export_excel

    def export_df_to_excel(self, sheet_name: str, start_row_ex: int = 0, start_col_ex: int = 0):
        """
        Exporta um Dataframe para um arquivo e planilha do Excel especificados.

        Args:
            df (pd.DataFrame): O DataFrame para exportar.
            sheet_name (str): O nome da planilha na qual a gravacao será feita.
            start_row (int):A linha para começar a escrever os dados.
            start_col (int): A coluna oara começar a escrever os dados

        Raises:
            Exception: Se houver um erro durenate o processo de gravação no excel.
        """
        try:
            with pd.ExcelWriter(path=self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                self.df_to_export_excel.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row_ex, startcol=start_col_ex)
            logger.info(f"DataFrame successfully exported to '{self.file_path}' sheet '{sheet_name}'.")
        except Exception as e:
            logger.error(f"Error exporting DataFrame to Excel: {e}")
            raise Exception(f"Failed to export DataFrame to Excel. Error: {e}") from e

    def format_excel_columns(self, sheet_name: str, date_columns: list[str], currency_columns: list[str]):
        """
        Aplica formatos numericos especificos a colunas em uma planilha excel.

        Args:
            sheet_name (str): O nome da planilha para formatar.
            date_columns (list[str]): Uma lista de letrars de colunas (ex., ['G', 'H']) para datas.
            currency_columns (list[str]): Uma lista de letrars de colunas(ex., ['J']) para formatar moedas.

        Raises:
            FileNotFoundError: Se o excel não existir
            KeyError: Se o nome especificado da planilha nao existir.
            Exception: Para outros erros durante a formatacao.
        """
        if not self.file_path.exists():
            logger.error(f"Excel file not found for formatting: {self.file_path}")
            raise FileNotFoundError(f"The file '{self.file_path}' does not exist.")
        
        wb = load_workbook(self.file_path)
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook.")
            raise KeyError(f"Sheet '{sheet_name}' does not exist in the Excel file.")
        
        try:
            ws = wb[sheet_name]
            for col_letter in date_columns:
                for cell in ws[col_letter][1:]: #pula o cabecalho
                    cell.number_format = "DD/MM/YYYY"

            for col_letter in currency_columns:
                for cell in ws[col_letter][1:]:  #pula o cabecalho
                    cell.number_format = r'#,##0.00; [Red](#,##0.00)'
            wb.save(self.file_path)
            logger.info(f"Excel formatting applied to '{sheet_name}'.")
        except Exception as e:
            logger.error(f"Error formatting Excel file '{self.file_path}': {e}")
            raise Exception(f"Failed to format Excel file. Error: {e}") from e

