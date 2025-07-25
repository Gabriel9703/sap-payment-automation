import logging
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd  # type: ignore
import pytest
from openpyxl import load_workbook  # type: ignore

from core.processing_data import ExcelFormater, ExcelProcessor

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

@pytest.fixture
def sample_excel_file():
    """Cria um arquivo Excel temporário para testes"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        file_path = Path(tmp.name)
    
    data = {
        'No Doc SAP': ['DOC001', 'DOC002', 'DOC003'],
        'Tipo Doc': ['Fatura', 'Contas a Pagar', 'Fatura'],
        'Conta': ['12345', '12345', '67890'],
        'Data Vencimento': ['01/01/2023', '15/01/2023', '31/12/2022'],
        'Valor': ['1.000,50', '2.500,00', '3.250,75']
    }
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        df.to_excel(writer, sheet_name='Sheet2', index=False)
    
    yield file_path
    
    # Limpeza após o teste
    if file_path.exists():
        file_path.unlink()

@pytest.fixture
def sample_dataframe():
    """Retorna um DataFrame de exemplo para testes"""
    data = {
        'No Doc SAP': ['DOC001', 'DOC002', 'DOC003'],
        'Tipo Doc': ['Fatura', 'Contas a Pagar', 'Fatura'],
        'Conta': ['12345', '12345', '67890'],
        'Data Vencimento': ['01/01/2023', '15/01/2023', '31/12/2022'],
        'Valor': ['1.000,50', '2.500,00', '3.250,75']
    }
    return pd.DataFrame(data)

@pytest.fixture
def excel_processor(sample_excel_file):
    """Fixture para ExcelProcessor com arquivo temporário"""
    return ExcelProcessor(sample_excel_file)

# Testes para ExcelProcessor
def test_load_excel_to_df_success(excel_processor):
    """Testa o carregamento bem-sucedido de um arquivo Excel"""
    df = excel_processor.load_excel_to_df(sheet_name='Sheet1')
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 3
    assert 'No Doc SAP' in df.columns

def test_load_excel_to_df_sheet_not_found(excel_processor):
    """Testa o erro quando a planilha não existe"""
    with pytest.raises(ValueError):
        excel_processor.load_excel_to_df(sheet_name='InvalidSheet')

def test_load_excel_to_df_file_not_found():
    """Testa o erro quando o arquivo não existe"""
    with pytest.raises(FileNotFoundError):
        processor = ExcelProcessor(Path('nonexistent_file.xlsx'))
        processor.load_excel_to_df()

def test_load_excel_to_df_with_columns(excel_processor):
    """Testa o carregamento com colunas específicas"""
    df = excel_processor.load_excel_to_df(columns_to_use=['No Doc SAP', 'Tipo Doc'])
    assert  isinstance(df, pd.DataFrame)
    assert len(df.columns) == 2
    assert 'No Doc SAP' in df.columns
    assert 'Tipo Doc' in df.columns

def test_view_df_success(excel_processor, sample_dataframe):
    """Testa a visualização de colunas específicas"""
    excel_processor.view_df(sample_dataframe, ['No Doc SAP', 'Conta'])

def test_view_df_column_not_found(excel_processor, sample_dataframe):
    """Testa o erro quando a coluna não existe"""
    with pytest.raises(KeyError):
        excel_processor.view_df(sample_dataframe, ['InvalidColumn'])

def test_filter_open_invoices_success(excel_processor, sample_dataframe):
    """Testa a filtragem de faturas em aberto"""
    filtered_df = excel_processor.filter_open_invoices(
        sample_dataframe, 
        type_doc_column='Tipo Doc',
        exclude_value='Contas a Pagar',
        account_column='Conta',
        account_code='12345'
    )
    assert isinstance(filtered_df, pd.DataFrame)
    assert len(filtered_df) == 1
    assert filtered_df.iloc[0]['No Doc SAP'] == 'DOC001' # type: ignore

def test_filter_open_invoices_column_not_found(excel_processor, sample_dataframe):
    """Testa o erro quando colunas necessárias não existem"""
    with pytest.raises(KeyError):
        excel_processor.filter_open_invoices(
            sample_dataframe,
            type_doc_column='InvalidColumn',
            exclude_value='Contas a Pagar',
            account_column='Conta',
            account_code='12345'
        )

def test_reconcile_invoices_success(excel_processor, sample_dataframe):
    """Testa a reconciliação de faturas"""
    paid_invoices = pd.DataFrame({'No Doc SAP': ['DOC002'], 'Status': ['Pago']})
    reconciled_df = excel_processor.reconcile_invoices(sample_dataframe, paid_invoices)
    assert isinstance(reconciled_df, pd.DataFrame)
    assert len(reconciled_df) == 2  # DOC001 e DOC003
    assert 'DOC002' not in reconciled_df['No Doc SAP'].values

def test_reconcile_invoices_missing_column(excel_processor, sample_dataframe):
    """Testa o erro quando a coluna 'No Doc SAP' está faltando"""
    paid_invoices = pd.DataFrame({'InvalidColumn': ['DOC002']})
    with pytest.raises(ValueError):
        excel_processor.reconcile_invoices(sample_dataframe, paid_invoices)

def test_reconcile_invoices_empty_df_raises(excel_processor, sample_dataframe):
    df_empty = pd.DataFrame()
    with pytest.raises(ValueError):
        excel_processor.reconcile_invoices(df_empty, sample_dataframe)

def test_convert_date_df_success(excel_processor, sample_dataframe):
    """Testa a conversão de colunas de data"""
    converted_df = excel_processor.convert_date_df(sample_dataframe, ['Data Vencimento'])
    assert isinstance(converted_df, pd.DataFrame)
    assert pd.api.types.is_datetime64_any_dtype(converted_df['Data Vencimento'])

def test_convert_date_df_column_not_found(excel_processor, sample_dataframe):
    """Testa o erro quando a coluna de data não existe"""
    with pytest.raises(KeyError):
        excel_processor.convert_date_df(sample_dataframe, ['InvalidDateColumn'])

def test_convert_values_df_success(excel_processor, sample_dataframe):
    """Testa a conversão de valores numéricos"""
    converted_df = excel_processor.convert_values_df(sample_dataframe, 'Valor')
    assert isinstance(converted_df, pd.DataFrame)
    assert pd.api.types.is_numeric_dtype(converted_df['Valor'])

def test_convert_values_df_column_not_found(excel_processor, sample_dataframe):
    """Testa o comportamento quando a coluna de valor não existe"""
    converted_df = excel_processor.convert_values_df(sample_dataframe, 'InvalidColumn')
    assert isinstance(converted_df, pd.DataFrame)

def test_filter_by_due_date_success(excel_processor, sample_dataframe):
    """Testa a filtragem por data de vencimento"""
    # Primeiro converter a coluna de data
    df = excel_processor.convert_date_df(sample_dataframe, ['Data Vencimento'])
    filtered_df = excel_processor.filter_by_due_date(df, 'Data Vencimento', '15/01/2023')
    assert isinstance(filtered_df, pd.DataFrame)
    assert len(filtered_df) == 3  
    assert filtered_df['Data Vencimento'].iloc[0] <= datetime(2023, 1, 15)

def test_filter_by_due_date_invalid_date(excel_processor, sample_dataframe):
    """Testa o erro com formato de data inválido"""
    df = excel_processor.convert_date_df(sample_dataframe, ['Data Vencimento'])
    with pytest.raises(ValueError):
        excel_processor.filter_by_due_date(df, 'Data Vencimento', 'invalid-date')

def test_filter_by_due_date_missing_column(excel_processor, sample_dataframe):
    """Testa o erro quando a coluna de data está faltando"""
    with pytest.raises(ValueError):
        excel_processor.filter_by_due_date(sample_dataframe, 'InvalidColumn', '01/01/2023')

def test_concatenate_dfs_success(excel_processor, sample_dataframe):
    """Testa a concatenação de DataFrames"""
    df2 = pd.DataFrame({'No Doc SAP': ['DOC004'], 'Tipo Doc': ['Fatura'], 'Conta': ['12345']})
    concatenated_df = excel_processor.concatenate_dfs(sample_dataframe, df2)
    assert isinstance(concatenated_df, pd.DataFrame)
    assert len(concatenated_df) == 4
    assert 'DOC004' in concatenated_df['No Doc SAP'].values
 

# -----------------------------  Testes para ExcelFormater  ------------------------------
@pytest.fixture
def excel_formater(sample_excel_file, sample_dataframe):
    """Fixture para ExcelFormater com arquivo temporário"""
    return ExcelFormater(sample_excel_file, sample_dataframe)

def test_export_df_to_excel_success(excel_formater):
    """Testa a exportação bem-sucedida para Excel"""
    # Teste de escrita em uma nova planilha
    excel_formater.export_df_to_excel(sheet_name='TestSheet')
    
    # Verificar se o arquivo foi criado e contém a planilha
    wb = load_workbook(excel_formater.file_path)
    assert 'TestSheet' in wb.sheetnames
    
    # Teste de escrita em posição específica
    excel_formater.export_df_to_excel(sheet_name='TestSheet2', start_row_ex=5, start_col_ex=2)
    wb = load_workbook(excel_formater.file_path)
    ws = wb['TestSheet2']
    assert ws.cell(row=6, column=3).value == 'No Doc SAP'
    assert ws.cell(row=7, column=3).value == 'DOC001'  # Verifica a posição (row+1, col+1)

def test_export_df_to_excel_error():
    """Testa o erro durante a exportação para Excel"""
    # Tentar escrever em um arquivo inexistente deve falhar
    invalid_processor = ExcelFormater(Path('invalid_path.xlsx'), pd.DataFrame())
    with pytest.raises(Exception):
        invalid_processor.export_df_to_excel(sheet_name='TestSheet')

def test_format_excel_columns_success(excel_formater):
    """Testa a formatação de colunas no Excel"""
    # Primeiro exportar os dados
    excel_formater.export_df_to_excel(sheet_name='FormattedSheet')
    
    # Aplicar formatação
    excel_formater.format_excel_columns(
        sheet_name='FormattedSheet',
        date_columns=['D'],  
        currency_columns=['E']  
    )
    
    # Verificar a formatação
    wb = load_workbook(excel_formater.file_path)
    ws = wb['FormattedSheet']
    
    # Verificar formato de data
    assert ws['D2'].number_format == "DD/MM/YYYY"
    
    # Verificar formato de moeda
    assert ws['E2'].number_format == r'#,##0.00; [Red](#,##0.00)'

def test_format_excel_columns_sheet_not_found(excel_formater):
    """Testa o erro quando a planilha não existe"""
    with pytest.raises(KeyError):
        excel_formater.format_excel_columns(
            sheet_name='NonexistentSheet',
            date_columns=['D'],
            currency_columns=['E']
        )

def test_format_excel_columns_file_not_found():
    """Testa o erro quando o arquivo Excel não existe"""
    formatter = ExcelFormater(Path('nonexistent_file.xlsx'), pd.DataFrame())
    with pytest.raises(FileNotFoundError):
        formatter.format_excel_columns(
            sheet_name='Sheet1',
            date_columns=['D'],
            currency_columns=['E']
        )